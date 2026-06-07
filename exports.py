"""
exports.py
All output: Excel workbooks and PNG line charts. Consumes the ranked
``VariableResult`` produced by ``evaluation.py``.

Public classes
──────────────
ForecastExporter      — single dependent-variable run (used by main.py)
BatchForecastExporter — multi-variable batch run (used by batch_run.py)

Notes
─────
- Models are ranked by **backtest MASE** (lower is better); the best is marked ★.
- The recent reported actuals are shown for reference only — they are not the ranking basis.
- Filenames carry NO timestamp, so outputs stay stable across re-runs: a re-run overwrites
  the prior workbook and charts of the same name.
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Dict, List, Optional

import matplotlib
matplotlib.use("Agg")
import matplotlib.dates as mdates
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd

from evaluation import ModelEval, VariableResult


# ─────────────────────────────────────────────────────────────────────────────
# Shared helpers
# ─────────────────────────────────────────────────────────────────────────────

def _safe_name(s: str) -> str:
    return re.sub(r"[^A-Za-z0-9._-]+", "_", str(s)).strip("_") or "var"


def _label(m: ModelEval) -> str:
    """Model display name with a leading ★ for the best and a ⚠ if any step is flagged."""
    star = "★ " if m.rank == 1 else ""
    warn = " ⚠" if any(f for f in m.flags) else ""
    return f"{star}{m.name}{warn}"


def _find_ci(vr: VariableResult):
    """Return (lower, upper) arrays from the best model that has a CI, else the
    first model that does (typically ARIMAX); (None, None) if none."""
    ordered = sorted(vr.models, key=lambda m: m.rank)
    for m in ordered:
        if m.final.ci_lower is not None and m.final.ci_upper is not None:
            return m.name, np.asarray(m.final.ci_lower), np.asarray(m.final.ci_upper)
    return None, None, None


def _forecasts_df(vr: VariableResult) -> pd.DataFrame:
    """Forecast table: one row per forecast period, one column per model (ranked)."""
    data: Dict[str, object] = {
        vr.time_col: list(vr.tail_index),
        f"{vr.dep_col} (reported)": [_round_forecast(v) for v in vr.tail_actuals.values],
    }
    for m in vr.models:
        data[f"{_label(m)}"] = [_round_forecast(v) for v in m.final.point]
    ci_name, lo, hi = _find_ci(vr)
    if lo is not None:
        data[f"{ci_name} CI Lower 95%"] = [_round_forecast(v) for v in lo]
        data[f"{ci_name} CI Upper 95%"] = [_round_forecast(v) for v in hi]
    best = vr.models[0]
    data["Implausible-jump flag"] = [f or "" for f in best.flags]
    return pd.DataFrame(data)


def _metrics_df(vr: VariableResult) -> pd.DataFrame:
    """Metrics table ranked by backtest MASE."""
    rows = []
    for m in vr.models:
        meta = m.meta or {}
        order = meta.get("order")
        sorder = meta.get("seasonal_order")
        order_str = "—"
        if order is not None:
            order_str = f"{tuple(order)}"
            if sorder and any(sorder[:3]):
                order_str += f"x{tuple(sorder)}"
        rows.append({
            "Rank": m.rank,
            "Model": _label(m),
            "MASE": _round(m.metrics.get("mase"), 3),
            "sMAPE (%)": _round(m.metrics.get("smape"), 2),
            "RMSE": _round(m.metrics.get("rmse"), 4),
            "MAE": _round(m.metrics.get("mae"), 4),
            "MAPE (%)": _round(m.metrics.get("mape"), 2),
            "Order": order_str,
            "AIC": _round(meta.get("aic"), 2) if isinstance(meta.get("aic"), (int, float)) else "—",
            "Beats naive": "no" if (m.metrics.get("mase") or np.nan) >= 1 else "yes",
            "Flagged": "yes" if any(f for f in m.flags) else "",
        })
    return pd.DataFrame(rows)


def _diagnostics_df(vr: VariableResult) -> pd.DataFrame:
    """Variable metadata + ARIMAX residual diagnostics + run notes."""
    rep = vr.selection_report or {}
    y_order = rep.get("_y_order")

    def _exog_line(c, d):
        s = f"{c}: I({d['order']}) r={d.get('spearman')}"
        if d.get("coint_p") is not None:
            s += f" coint_p={d['coint_p']}"
        return s + f" [{d['basis']}]"

    exog_orders = "; ".join(
        _exog_line(c, d) for c, d in rep.items() if c != "_y_order"
    ) or "—"
    rows = [
        {"Item": "Dependent variable", "Value": vr.dep_col},
        {"Item": "Seasonal period (m)", "Value": vr.period},
        {"Item": "Dependent integration order", "Value": f"I({y_order})" if y_order is not None else "—"},
        {"Item": "Selected exogenous", "Value": ", ".join(vr.selected_exog) or "none"},
        {"Item": "Exog: order, relevance (r on stationary basis), keep-basis", "Value": exog_orders},
        {"Item": "Ranking basis", "Value": "rolling-origin backtest MASE on settled history"},
        {"Item": "Recent actuals", "Value": "shown for reference; excluded from ranking (reporting lag)"},
        {"Item": "Implausible-jump flag", "Value": "non-empty only when a forecast step jumps "
            ">3× the largest historical move or leaves the historical range; blank = no concern"},
    ]
    arimax = next((m for m in vr.models if m.name == "ARIMAX"), None)
    if arimax is not None:
        meta = arimax.meta or {}
        diag = meta.get("diagnostics", {}) or {}
        lam = meta.get("exog_lambdas") or {}
        lam_str = ", ".join(f"{k}: λ={v}" for k, v in lam.items()) or "—"
        rows += [
            {"Item": "ARIMAX transform (y)", "Value": meta.get("transform", "—")},
            {"Item": "ARIMAX exog linearization (Yeo-Johnson λ; 1=identity)", "Value": lam_str},
            {"Item": "ARIMAX Ljung-Box p (autocorr)", "Value": _round(diag.get("ljung_box_p"), 4)},
            {"Item": "ARIMAX Shapiro p (normality)", "Value": _round(diag.get("normality_p"), 4)},
            {"Item": "ARIMAX ARCH p (heterosked.)", "Value": _round(diag.get("arch_p"), 4)},
            {"Item": "ARIMAX residual notes", "Value": diag.get("notes", "—")},
        ]
    return pd.DataFrame(rows)


def _round(v, n):
    try:
        if v is None or (isinstance(v, float) and not np.isfinite(v)):
            return "—"
        return round(float(v), n)
    except Exception:
        return "—"


def _round_forecast(v):
    """Display rounding for forecast/actual values: drop the decimals when they account
    for less than 0.1% of the value (e.g. 123456.12 → 123456), otherwise keep two
    decimals (e.g. 1.12 → 1.12)."""
    try:
        f = float(v)
    except (TypeError, ValueError):
        return v
    if not np.isfinite(f) or f == 0:
        return f
    frac = abs(f - round(f))
    if frac / abs(f) < 0.001:
        return int(round(f))
    return round(f, 2)


def _autofit(ws) -> None:
    for col_cells in ws.columns:
        max_len = max((len(str(c.value)) for c in col_cells if c.value is not None), default=10)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 60)


def _draw_chart(ax, vr: VariableResult) -> None:
    use_dates = pd.api.types.is_datetime64_any_dtype(pd.Index(vr.tail_index))

    train = vr.train_actuals
    ax.plot(train.index, train.values, color="#2c7bb6", linewidth=1.5, label="Settled history")

    ax.plot(vr.tail_index, vr.tail_actuals.values, "o--", color="#7f7f7f",
            linewidth=1.2, markersize=5, label="Reported")

    palette = ["#d7191c", "#1a9641", "#f4a261", "#7b2d8b", "#00bcd4", "#2ca02c", "#e377c2", "#8c564b"]
    for m, color in zip(vr.models, palette):
        is_best = m.rank == 1
        ax.plot(
            vr.tail_index, m.final.point,
            ("s-" if is_best else "^--"), color=color,
            linewidth=2.0 if is_best else 1.3, markersize=6 if is_best else 4,
            label=f"{_label(m)} (MASE={m.metrics.get('mase', float('nan')):.3f})",
            zorder=5 if is_best else 3,
        )

    ci_name, lo, hi = _find_ci(vr)
    if lo is not None:
        ax.fill_between(vr.tail_index, lo, hi, alpha=0.12, color="#d7191c",
                        label=f"{ci_name} 95% CI")

    if len(train) > 0:
        ax.axvline(train.index[-1], color="gray", linestyle=":", linewidth=1, alpha=0.6)
    if use_dates:
        ax.xaxis.set_major_formatter(mdates.AutoDateFormatter(mdates.AutoDateLocator()))

    best = vr.models[0]
    ax.set_title(f"Forecast — {vr.dep_col}  |  Best: ★ {best.name}  (MASE={best.metrics.get('mase', float('nan')):.3f})",
                 fontsize=13, fontweight="bold")
    ax.set_xlabel(vr.time_col, fontsize=11)
    ax.set_ylabel(vr.dep_col, fontsize=11)
    ax.legend(loc="best", fontsize=8, framealpha=0.85)
    ax.grid(True, alpha=0.25)


# ─────────────────────────────────────────────────────────────────────────────
# Single-run exporter (used by main.py)
# ─────────────────────────────────────────────────────────────────────────────

class ForecastExporter:
    """
    Writes (stable names, no timestamp, so a re-run overwrites the prior output):
      • "Estimates {dep}.xlsx"  — Forecasts / Metrics / Diagnostics
      • "Estimates {dep}.png"   — line chart
    """

    def __init__(self, output_dir: str = ".") -> None:
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def excel_path(self, dep_col: str) -> Path:
        return self.output_dir / f"Estimates {_safe_name(dep_col)}.xlsx"

    def chart_path(self, dep_col: str) -> Path:
        return self.output_dir / f"Estimates {_safe_name(dep_col)}.png"

    def export_excel(self, vr: VariableResult) -> Path:
        path = self.excel_path(vr.dep_col)
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            _forecasts_df(vr).to_excel(writer, sheet_name="Forecasts", index=False)
            _autofit(writer.sheets["Forecasts"])
            _metrics_df(vr).to_excel(writer, sheet_name="Metrics", index=False)
            _autofit(writer.sheets["Metrics"])
            _diagnostics_df(vr).to_excel(writer, sheet_name="Diagnostics", index=False)
            _autofit(writer.sheets["Diagnostics"])
        print(f"  Excel saved  →  {path}")
        return path

    def export_chart(self, vr: VariableResult) -> Path:
        use_dates = pd.api.types.is_datetime64_any_dtype(pd.Index(vr.tail_index))
        fig, ax = plt.subplots(figsize=(14, 6))
        _draw_chart(ax, vr)
        if use_dates:
            fig.autofmt_xdate()
        plt.tight_layout()
        path = self.chart_path(vr.dep_col)
        fig.savefig(path, dpi=300, bbox_inches="tight")
        plt.close(fig)
        print(f"  Chart saved  →  {path}")
        return path


# ─────────────────────────────────────────────────────────────────────────────
# Batch exporter (used by batch_run.py)
# ─────────────────────────────────────────────────────────────────────────────

class BatchForecastExporter:
    """
    Accumulates ``VariableResult``s, then writes (stable names, no timestamp, so a re-run
    overwrites the prior output):
      Excel  →  {prefix}.xlsx       (Master Forecasts, Master Metrics, per-var sheets)
      Charts →  {prefix}_{var}.png  (one per variable)
    """

    def __init__(self, output_dir: str = ".", prefix: str = "batch_forecast") -> None:
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.prefix = prefix
        self._runs: List[VariableResult] = []

    @property
    def has_results(self) -> bool:
        return len(self._runs) > 0

    @property
    def excel_path(self) -> Path:
        return self.output_dir / f"{self.prefix}.xlsx"

    def chart_path(self, dep_col: str) -> Path:
        return self.output_dir / f"{self.prefix}_{_safe_name(dep_col)}.png"

    def add_run(self, vr: VariableResult) -> None:
        self._runs.append(vr)

    def export_excel(self) -> Path:
        master_fc: List[Dict] = []
        master_met: List[Dict] = []
        with pd.ExcelWriter(self.excel_path, engine="openpyxl") as writer:
            for vr in self._runs:
                best = vr.models[0]
                for i, idx_val in enumerate(vr.tail_index):
                    row: Dict = {
                        "variable": vr.dep_col,
                        vr.time_col: idx_val,
                        "reported": _round_forecast(vr.tail_actuals.iloc[i]),
                        "best_model": best.name,
                        "best_forecast": _round_forecast(best.final.point[i]),
                        "implausible_jump_flag": best.flags[i] if i < len(best.flags) else "",
                    }
                    for m in vr.models:
                        row[m.name] = _round_forecast(m.final.point[i])
                    master_fc.append(row)

                met = _metrics_df(vr)
                met.insert(0, "variable", vr.dep_col)
                master_met.append(met)

                # per-variable sheet: forecasts, then diagnostics below
                sheet = f"FC — {vr.dep_col}"[:31]
                fdf = _forecasts_df(vr)
                fdf.to_excel(writer, sheet_name=sheet, index=False)
                ddf = _diagnostics_df(vr)
                ddf.to_excel(writer, sheet_name=sheet, index=False, startrow=len(fdf) + 3)
                _autofit(writer.sheets[sheet])

            if master_fc:
                pd.DataFrame(master_fc).to_excel(writer, sheet_name="Master Forecasts", index=False)
                _autofit(writer.sheets["Master Forecasts"])
            if master_met:
                pd.concat(master_met, ignore_index=True).to_excel(
                    writer, sheet_name="Master Metrics", index=False)
                _autofit(writer.sheets["Master Metrics"])

        # reorder so master sheets come first (cosmetic)
        try:
            import openpyxl
            wb = openpyxl.load_workbook(self.excel_path)
            first = [n for n in ("Master Forecasts", "Master Metrics") if n in wb.sheetnames]
            rest = [n for n in wb.sheetnames if n not in first]
            wb._sheets = [wb[n] for n in first + rest]
            wb.save(self.excel_path)
        except Exception:
            pass

        print(f"  Excel saved  →  {self.excel_path}")
        return self.excel_path

    def export_charts(self) -> List[Path]:
        paths: List[Path] = []
        for vr in self._runs:
            use_dates = pd.api.types.is_datetime64_any_dtype(pd.Index(vr.tail_index))
            fig, ax = plt.subplots(figsize=(14, 6))
            _draw_chart(ax, vr)
            if use_dates:
                fig.autofmt_xdate()
            plt.tight_layout()
            p = self.chart_path(vr.dep_col)
            fig.savefig(p, dpi=300, bbox_inches="tight")
            plt.close(fig)
            print(f"  Chart saved  →  {p}")
            paths.append(p)
        return paths
